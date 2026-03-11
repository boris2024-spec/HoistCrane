from rest_framework import viewsets, filters, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django_filters import rest_framework as django_filters
from .models import Equipment, EquipmentSpecification
from .serializers import EquipmentSerializer, EquipmentListSerializer
from core.permissions import CanModifyEquipment
import csv
import io
from datetime import datetime
import chardet
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from django.db import transaction
import uuid
import re


def _normalize_header(value: str) -> str:
    if value is None:
        return ''
    value = str(value)
    value = value.replace('\ufeff', '')  # BOM
    value = value.replace('\u00a0', ' ')  # NBSP
    value = re.sub(r"\s+", " ", value).strip()
    return value


def _get_first(row_dict: dict, candidates: list[str]) -> str:
    for key in candidates:
        if key in row_dict and row_dict.get(key) not in (None, ''):
            return str(row_dict.get(key)).strip()
    return ''


def _parse_date(value: str):
    if not value:
        return None
    value = str(value).strip()
    if value == '':
        return None

    # Try to extract a date token from a longer string
    m = re.search(
        r"(\d{1,2}[\./-]\d{1,2}[\./-]\d{2,4}|\d{4}-\d{2}-\d{2})", value)
    token = m.group(1) if m else value

    for fmt in ['%d/%m/%Y', '%d.%m.%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y', '%d.%m.%y']:
        try:
            return datetime.strptime(token.strip(), fmt).date()
        except Exception:
            continue
    return None


def _map_equipment_type(raw: str) -> str:
    raw = (raw or '').strip()
    if raw == '':
        return 'lifting_facilities'

    # Valid model choices
    valid_choices = {'lifting_accessories', 'no_inspection_required',
                     'forklifts', 'lifting_facilities'}

    # If already a valid DB value, return as-is
    raw_lower = raw.lower()
    if raw_lower in valid_choices:
        return raw_lower

    mapping = {
        # Hebrew labels → DB keys
        'אביזרי הרמה': 'lifting_accessories',
        'אביזר הרמה': 'lifting_accessories',
        'לא חייב בבדיקה': 'no_inspection_required',
        'לא חייב': 'no_inspection_required',
        'מלגזות': 'forklifts',
        'מלגזה': 'forklifts',
        'forklift': 'forklifts',
        'מתקני הרמה': 'lifting_facilities',
        'מתקן הרמה': 'lifting_facilities',
        'מנוף': 'lifting_facilities',
        'מנוף הרמה': 'lifting_facilities',
        'מנופון': 'lifting_facilities',
        'crane': 'lifting_facilities',
        'hoist': 'lifting_facilities',
        'עגורן': 'lifting_facilities',
        'מעלית': 'lifting_facilities',
        'elevator': 'lifting_facilities',
        'במה': 'lifting_facilities',
        'platform': 'lifting_facilities',
        'genie': 'lifting_facilities',
        'lift': 'lifting_facilities',
        'ליפט': 'lifting_facilities',
    }

    for token, mapped in mapping.items():
        if token in raw_lower or token in raw:
            return mapped
    return 'lifting_facilities'


def _map_status(raw: str) -> str:
    raw = (raw or '').strip()
    if raw == '':
        return 'active'

    raw_lower = raw.lower()
    if raw_lower in {'active', 'maintenance', 'inactive', 'retired'}:
        return raw_lower

    he_map = {
        'פעיל': 'active',
        'בתוקף': 'active',
        'תקף': 'active',
        'תחזוקה': 'maintenance',
        'בתחזוקה': 'maintenance',
        'לא פעיל': 'inactive',
        'לא בתוקף': 'inactive',
        'לא תקף': 'inactive',
        'מושבת': 'inactive',
        'הוצא משימוש': 'retired',
        'גרוטאה': 'retired',
    }
    for token, mapped in he_map.items():
        if token in raw:
            return mapped

    return 'active'


def _extract_two_dates_from_text(text: str):
    if not text:
        return (None, None)
    tokens = re.findall(
        r"\d{1,2}[\./-]\d{1,2}[\./-]\d{2,4}|\d{4}-\d{2}-\d{2}", str(text))
    parsed = [_parse_date(t) for t in tokens]
    parsed = [d for d in parsed if d is not None]
    if len(parsed) >= 2:
        return (parsed[0], parsed[1])
    if len(parsed) == 1:
        return (parsed[0], None)
    return (None, None)


HEADER_CANDIDATES = {
    'equipment_number': ['מספר סידורי פנימי', 'מספר ציוד', 'פריט ציוד', 'Equipment Number', 'Internal Serial'],
    'equipment_type': ['תחום ציוד', 'סוג ציוד', 'סוג', 'פריט ציוד', 'Equipment Type'],
    'super_domain': ['תחום על', 'Super Domain'],
    'status': ['סטטוס פריט ציוד', 'סטטוס', 'סטטוס פרט ציוד', 'Status'],
    'inspection_status': ['סטטוס בדיקות', 'סטטוס בדיקה', 'Inspection Status'],
    'internal_serial_number': ['מספר סידורי פנימי', 'Internal Serial'],
    'company': ['חברה', 'מעסיק', 'מזמין', 'Company', 'Employer'],
    'service_company': ['חברת שירות / קבלן', 'חברת שירות', 'Service Company'],
    'wing': ['אגף', 'Wing'],
    'division': ['חטיבה', 'Division'],
    'department': ['מחלקה', 'Department'],
    'sub_department': ['תת מחלקה', 'Sub Department'],
    'unit': ['יחידה', 'Unit'],
    'country': ['מדינה', 'Country'],
    'district': ['מחוז / איזור', 'מחוז', 'איזור', 'District'],
    'city': ['עיר / יישוב', 'עיר', 'יישוב', 'City'],
    'site_name': ['אתר / סניף', 'אתר', 'אתר / סרק', 'Site', 'Site/Branch'],
    'yam_number': ['מספר יא״מ', 'YAM Number'],
    'site_status': ['סטטוס אתר / סניף', 'סטטוס אתר', 'Site Status'],
    'campus': ['קמפוס', 'Campus'],
    'address': ['כתובת', 'Address'],
    'building': ['מבנה / מתקן', 'מבנה', 'מתקן', 'Building'],
    'floor_number': ['קומה', 'Floor'],
    'room': ['חדר', 'Room'],
    'workplace_name': ['מקום עבודה', 'Workplace'],
    'location': ['מיקום', 'מיקום פריט ציוד', 'Location'],
    'production_line': ['קו ייצור', 'Production Line'],
    'project': ['פרויקט', 'Project'],
    'license_number': ['מספר רישיון / רישוי', 'מספר רישיון', 'License Number'],
    'manufacturer': ['יצרן', 'Manufacturer'],
    'manufacture_date': ['תאריך ייצור', 'Manufacture Date'],
    'manufacture_year': ['שנת ייצור', 'Manufacture Year'],
    'serial_number': ['מספר סידורי יצרן', 'Serial Number', 'מסלקד'],
    'warranty_expiry': ['פקיעת תוקף אחריות', 'Warranty Expiry'],
    'model': ['דגם', 'Model'],
    'description': ['תאור', 'תיאור', 'Description'],
    'notes': ['הערה', 'הערות', 'Notes'],
    'tag': ['תגית', 'Tag'],
    'responsible': ['אחראי/ת', 'אחראי', 'בודק', 'בודק/ת', 'Inspector'],
    'periodic_inspections': ['בדיקות תקופתיות', 'בדיקה אחרונה', 'בדיקה הבאה', 'Inspection Dates'],
    'equipment_set': ['ערכת ציוד', 'Equipment Set'],
    'certified_workers': ['עובדים מוסמכים', 'Certified Workers'],
    'safe_working_load': ['עומס עבודה בטוח', 'Safe Working Load', 'SWL'],
    'max_allowed_pressure': ['לחץ מירבי מותר', 'Max Allowed Pressure'],
    'measurement_unit': ['יחידת מדידה', 'Measurement Unit'],
    'measurement_resolution': ['רזולוציית מדידה', 'רזולוצית מדידה', 'Measurement Resolution'],
    'measurement_range': ['טווח מדידה', 'Measurement Range'],
    'url': ['URL', 'קישור'],
    'file_count': ['מספר קבצים', 'File Count'],
    'image_url': ['תמונה', 'Image'],
}


def _import_normalized_rows(*, user, rows_iter):
    """Import rows that already have normalized headers.

    rows_iter yields tuples: (row_index:int, normalized_row:dict[str, Any]).
    """
    imported_count = 0
    skipped_existing = 0
    error_count = 0
    errors = []

    def get_value(normalized_row: dict, field_key: str) -> str:
        candidates = HEADER_CANDIDATES[field_key]
        normalized_candidates = [_normalize_header(c) for c in candidates]
        return _get_first(normalized_row, normalized_candidates)

    with transaction.atomic():
        for row_index, normalized_row in rows_iter:
            try:
                equipment_number = (
                    get_value(normalized_row, 'equipment_number') or '').strip()
                # Fallback: if equipment_number is empty, try internal_serial_number
                if not equipment_number:
                    equipment_number = (
                        get_value(normalized_row, 'internal_serial_number') or '').strip()
                if not equipment_number:
                    raise ValueError(
                        'Missing equipment number (פריט ציוד / מספר סידורי פנימי)')

                if Equipment.objects.filter(equipment_number=equipment_number).exists():
                    skipped_existing += 1
                    continue

                equipment_type_raw = get_value(
                    normalized_row, 'equipment_type')
                status_raw = get_value(normalized_row, 'status')
                inspection_status_raw = get_value(
                    normalized_row, 'inspection_status')
                company = get_value(normalized_row, 'company')
                department = get_value(normalized_row, 'department')
                site_name = get_value(normalized_row, 'site_name')
                workplace_name = get_value(normalized_row, 'workplace_name')
                location = get_value(normalized_row, 'location')
                manufacturer = get_value(normalized_row, 'manufacturer')
                model = get_value(normalized_row, 'model')
                serial_number = get_value(normalized_row, 'serial_number')
                description = get_value(normalized_row, 'description')
                responsible = get_value(normalized_row, 'responsible')

                manufacture_date = _parse_date(
                    get_value(normalized_row, 'manufacture_date'))
                manufacture_year_raw = get_value(
                    normalized_row, 'manufacture_year')
                manufacture_year = None
                if manufacture_year_raw:
                    try:
                        manufacture_year = int(
                            str(manufacture_year_raw).strip()[:4])
                    except Exception:
                        manufacture_year = None
                if not manufacture_year and manufacture_date:
                    manufacture_year = manufacture_date.year

                periodic_text = get_value(
                    normalized_row, 'periodic_inspections')

                last_inspection = _get_first(
                    normalized_row, [_normalize_header('בדיקה אחרונה')])
                next_inspection = _get_first(
                    normalized_row, [_normalize_header('בדיקה הבאה')])

                last_inspection_date = _parse_date(last_inspection)
                next_inspection_date = _parse_date(next_inspection)
                if not last_inspection_date and not next_inspection_date:
                    last_inspection_date, next_inspection_date = _extract_two_dates_from_text(
                        periodic_text)

                # Gather all new fields from row
                super_domain = get_value(
                    normalized_row, 'super_domain') if 'super_domain' in HEADER_CANDIDATES else ''
                internal_serial = get_value(
                    normalized_row, 'internal_serial_number') if 'internal_serial_number' in HEADER_CANDIDATES else ''
                service_company_val = get_value(
                    normalized_row, 'service_company') if 'service_company' in HEADER_CANDIDATES else ''
                wing_val = get_value(
                    normalized_row, 'wing') if 'wing' in HEADER_CANDIDATES else ''
                division_val = get_value(
                    normalized_row, 'division') if 'division' in HEADER_CANDIDATES else ''
                sub_department_val = get_value(
                    normalized_row, 'sub_department') if 'sub_department' in HEADER_CANDIDATES else ''
                unit_val = get_value(
                    normalized_row, 'unit') if 'unit' in HEADER_CANDIDATES else ''
                country_val = get_value(
                    normalized_row, 'country') if 'country' in HEADER_CANDIDATES else ''
                district_val = get_value(
                    normalized_row, 'district') if 'district' in HEADER_CANDIDATES else ''
                city_val = get_value(
                    normalized_row, 'city') if 'city' in HEADER_CANDIDATES else ''
                yam_number_val = get_value(
                    normalized_row, 'yam_number') if 'yam_number' in HEADER_CANDIDATES else ''
                site_status_val = get_value(
                    normalized_row, 'site_status') if 'site_status' in HEADER_CANDIDATES else ''
                campus_val = get_value(
                    normalized_row, 'campus') if 'campus' in HEADER_CANDIDATES else ''
                address_val = get_value(
                    normalized_row, 'address') if 'address' in HEADER_CANDIDATES else ''
                building_val = get_value(
                    normalized_row, 'building') if 'building' in HEADER_CANDIDATES else ''
                floor_val = get_value(
                    normalized_row, 'floor_number') if 'floor_number' in HEADER_CANDIDATES else ''
                room_val = get_value(
                    normalized_row, 'room') if 'room' in HEADER_CANDIDATES else ''
                production_line_val = get_value(
                    normalized_row, 'production_line') if 'production_line' in HEADER_CANDIDATES else ''
                project_val = get_value(
                    normalized_row, 'project') if 'project' in HEADER_CANDIDATES else ''
                license_number_val = get_value(
                    normalized_row, 'license_number') if 'license_number' in HEADER_CANDIDATES else ''
                warranty_expiry_val = _parse_date(get_value(
                    normalized_row, 'warranty_expiry')) if 'warranty_expiry' in HEADER_CANDIDATES else None
                notes_val = get_value(
                    normalized_row, 'notes') if 'notes' in HEADER_CANDIDATES else ''
                tag_val = get_value(
                    normalized_row, 'tag') if 'tag' in HEADER_CANDIDATES else ''
                equipment_set_val = get_value(
                    normalized_row, 'equipment_set') if 'equipment_set' in HEADER_CANDIDATES else ''
                certified_workers_val = get_value(
                    normalized_row, 'certified_workers') if 'certified_workers' in HEADER_CANDIDATES else ''
                safe_working_load_val = get_value(
                    normalized_row, 'safe_working_load') if 'safe_working_load' in HEADER_CANDIDATES else ''
                max_pressure_val = get_value(
                    normalized_row, 'max_allowed_pressure') if 'max_allowed_pressure' in HEADER_CANDIDATES else ''
                measurement_unit_val = get_value(
                    normalized_row, 'measurement_unit') if 'measurement_unit' in HEADER_CANDIDATES else ''
                measurement_resolution_val = get_value(
                    normalized_row, 'measurement_resolution') if 'measurement_resolution' in HEADER_CANDIDATES else ''
                measurement_range_val = get_value(
                    normalized_row, 'measurement_range') if 'measurement_range' in HEADER_CANDIDATES else ''
                url_val = get_value(
                    normalized_row, 'url') if 'url' in HEADER_CANDIDATES else ''

                equipment = Equipment.objects.create(
                    equipment_number=equipment_number,
                    equipment_type=_map_equipment_type(equipment_type_raw),
                    super_domain=(super_domain or '').strip(),
                    status=_map_status(status_raw),
                    inspection_status=(inspection_status_raw or '').strip()[
                        :20] if inspection_status_raw else 'none',
                    internal_serial_number=(internal_serial or '').strip(),
                    manufacturer=(manufacturer or '').strip(),
                    model=(model or '').strip(),
                    serial_number=(serial_number or '').strip() or None,
                    manufacture_date=manufacture_date,
                    manufacture_year=manufacture_year,
                    license_number=(license_number_val or '').strip(),
                    warranty_expiry=warranty_expiry_val,
                    site_name=(site_name or '').strip(),
                    workplace_name=(workplace_name or '').strip(),
                    employer=(company or '').strip(),
                    service_company=(service_company_val or '').strip(),
                    wing=(wing_val or '').strip(),
                    division=(division_val or '').strip(),
                    department=(department or '').strip(),
                    sub_department=(sub_department_val or '').strip(),
                    unit=(unit_val or '').strip(),
                    country=(country_val or '').strip(),
                    district=(district_val or '').strip(),
                    city=(city_val or '').strip(),
                    yam_number=(yam_number_val or '').strip(),
                    site_status=(site_status_val or '').strip(),
                    campus=(campus_val or '').strip(),
                    address=(address_val or '').strip(),
                    building=(building_val or '').strip(),
                    floor_number=(floor_val or '').strip(),
                    room=(room_val or '').strip(),
                    location_details=(location or '').strip(),
                    production_line=(production_line_val or '').strip(),
                    project=(project_val or '').strip(),
                    description=(description or '').strip(),
                    notes=(notes_val or '').strip(),
                    tag=(tag_val or '').strip(),
                    inspector_name=(responsible or '').strip(),
                    periodic_inspections=(periodic_text or '').strip(),
                    equipment_set=(equipment_set_val or '').strip(),
                    certified_workers=(certified_workers_val or '').strip(),
                    safe_working_load=(safe_working_load_val or '').strip(),
                    max_allowed_pressure=(max_pressure_val or '').strip(),
                    measurement_unit=(measurement_unit_val or '').strip(),
                    measurement_resolution=(
                        measurement_resolution_val or '').strip(),
                    measurement_range=(measurement_range_val or '').strip(),
                    url=(url_val or '').strip(),
                    last_inspection_date=last_inspection_date,
                    next_inspection_date=next_inspection_date,
                    created_by=user,
                    updated_by=user,
                )

                imported_count += 1

            except Exception as e:
                error_count += 1
                errors.append(f"Row {row_index}: {str(e)}")

    return {
        'success': True,
        'imported': imported_count,
        'skipped_existing': skipped_existing,
        'errors': error_count,
        'error_details': errors[:10],
    }


class EquipmentFilter(django_filters.FilterSet):
    # Multiple choice filters
    equipment_type = django_filters.CharFilter(method='filter_multiple')
    status = django_filters.CharFilter(method='filter_status_multiple')
    inspection_status = django_filters.CharFilter(method='filter_multiple')

    # Text filters
    manufacturer = django_filters.CharFilter(lookup_expr='icontains')
    model = django_filters.CharFilter(lookup_expr='icontains')
    serial_number = django_filters.CharFilter(lookup_expr='icontains')
    internal_serial_number = django_filters.CharFilter(lookup_expr='icontains')
    site_name = django_filters.CharFilter(lookup_expr='icontains')
    inspector_name = django_filters.CharFilter(lookup_expr='icontains')
    employer = django_filters.CharFilter(lookup_expr='icontains')
    service_company = django_filters.CharFilter(lookup_expr='icontains')
    department = django_filters.CharFilter(lookup_expr='icontains')
    sub_department = django_filters.CharFilter(lookup_expr='icontains')
    division = django_filters.CharFilter(lookup_expr='icontains')
    wing = django_filters.CharFilter(lookup_expr='icontains')
    unit = django_filters.CharFilter(lookup_expr='icontains')
    workplace_name = django_filters.CharFilter(lookup_expr='icontains')
    country = django_filters.CharFilter(lookup_expr='icontains')
    district = django_filters.CharFilter(lookup_expr='icontains')
    city = django_filters.CharFilter(lookup_expr='icontains')
    campus = django_filters.CharFilter(lookup_expr='icontains')
    building = django_filters.CharFilter(lookup_expr='icontains')
    project = django_filters.CharFilter(lookup_expr='icontains')
    tag = django_filters.CharFilter(lookup_expr='icontains')
    super_domain = django_filters.CharFilter(lookup_expr='icontains')

    # Range filters
    capacity_min = django_filters.NumberFilter(
        field_name='capacity', lookup_expr='gte')
    capacity_max = django_filters.NumberFilter(
        field_name='capacity', lookup_expr='lte')
    height_min = django_filters.NumberFilter(
        field_name='height', lookup_expr='gte')
    height_max = django_filters.NumberFilter(
        field_name='height', lookup_expr='lte')
    manufacture_year_min = django_filters.NumberFilter(
        field_name='manufacture_year', lookup_expr='gte')
    manufacture_year_max = django_filters.NumberFilter(
        field_name='manufacture_year', lookup_expr='lte')

    # Date range filters
    last_inspection_date_from = django_filters.DateFilter(
        field_name='last_inspection_date', lookup_expr='gte')
    last_inspection_date_to = django_filters.DateFilter(
        field_name='last_inspection_date', lookup_expr='lte')
    next_inspection_date_from = django_filters.DateFilter(
        field_name='next_inspection_date', lookup_expr='gte')
    next_inspection_date_to = django_filters.DateFilter(
        field_name='next_inspection_date', lookup_expr='lte')

    def filter_multiple(self, queryset, name, value):
        """Handle multiple comma-separated values"""
        if value:
            values = [v.strip() for v in value.split(',')]
            return queryset.filter(**{f'{name}__in': values})
        return queryset

    def filter_status_multiple(self, queryset, name, value):
        """Handle multiple status values"""
        if value:
            values = [v.strip() for v in value.split(',')]
            return queryset.filter(status__in=values)
        return queryset

    class Meta:
        model = Equipment
        fields = [
            'equipment_type', 'super_domain', 'status', 'inspection_status',
            'manufacturer', 'model', 'serial_number', 'internal_serial_number',
            'site_name', 'inspector_name', 'employer', 'service_company',
            'department', 'sub_department', 'division', 'wing', 'unit',
            'workplace_name', 'country', 'district', 'city', 'campus',
            'building', 'project', 'tag',
            'capacity_min', 'capacity_max', 'height_min', 'height_max',
            'manufacture_year_min', 'manufacture_year_max',
            'last_inspection_date_from', 'last_inspection_date_to',
            'next_inspection_date_from', 'next_inspection_date_to'
        ]


class EquipmentViewSet(viewsets.ModelViewSet):
    queryset = Equipment.objects.select_related(
        'created_by', 'updated_by').all()
    permission_classes = [permissions.IsAuthenticated, CanModifyEquipment]
    filter_backends = [DjangoFilterBackend,
                       filters.SearchFilter, filters.OrderingFilter]
    filterset_class = EquipmentFilter
    search_fields = ['equipment_number', 'serial_number', 'internal_serial_number',
                     'manufacturer', 'model', 'site_name',
                     'workplace_name', 'employer', 'service_company',
                     'department', 'sub_department', 'division', 'wing', 'unit',
                     'country', 'district', 'city', 'campus', 'building',
                     'location_details', 'description', 'notes', 'tag',
                     'project', 'equipment_set', 'address']
    ordering_fields = ['created_at',
                       'equipment_number', 'next_inspection_date']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return EquipmentListSerializer
        return EquipmentSerializer

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user,
                        updated_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    @action(detail=False, methods=['get'], url_path='options')
    def options(self, request):
        """Return distinct filter option values for Equipment list UI."""
        inspector_names_qs = (
            Equipment.objects
            .exclude(inspector_name__isnull=True)
            .exclude(inspector_name__exact='')
            .values_list('inspector_name', flat=True)
            .distinct()
            .order_by('inspector_name')
        )

        return Response(
            {
                'inspector_names': list(inspector_names_qs),
            }
        )

    @action(detail=False, methods=['post'], url_path='bulk-delete')
    def bulk_delete(self, request):
        """Delete multiple equipment items by ids.

        Expected payload: {"ids": ["uuid1", "uuid2", ...]}
        """
        ids = request.data.get('ids', [])

        if isinstance(ids, str):
            ids = [part.strip() for part in ids.split(',') if part.strip()]

        if not isinstance(ids, (list, tuple)) or len(ids) == 0:
            return Response(
                {'detail': 'ids must be a non-empty list of UUIDs'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            id_list = sorted({str(uuid.UUID(str(value))) for value in ids})
        except (TypeError, ValueError, AttributeError):
            return Response(
                {'detail': 'ids must be a non-empty list of UUIDs'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        queryset = Equipment.objects.filter(id__in=id_list)
        found_ids = {str(value)
                     for value in queryset.values_list('id', flat=True)}
        not_found = [
            equipment_id for equipment_id in id_list if equipment_id not in found_ids]

        equipment_to_delete = queryset.count()

        with transaction.atomic():
            queryset.delete()

        return Response(
            {
                'requested': len(id_list),
                'deleted': equipment_to_delete,
                'not_found': not_found,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=['get'])
    def inspections(self, request, pk=None):
        """Get all inspections for this equipment"""
        equipment = self.get_object()
        inspections = equipment.inspections.all()
        from inspections.serializers import InspectionSerializer
        serializer = InspectionSerializer(inspections, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def documents(self, request, pk=None):
        """Get all documents for this equipment"""
        equipment = self.get_object()
        documents = equipment.documents.all()
        from documents.serializers import DocumentSerializer
        serializer = DocumentSerializer(documents, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def issues(self, request, pk=None):
        """Get all issues for this equipment"""
        equipment = self.get_object()
        issues = equipment.issues.all()
        from issues.serializers import IssueSerializer
        serializer = IssueSerializer(issues, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'], url_path='generate-pdf')
    def generate_pdf(self, request, pk=None):
        """Generate a PDF equipment card for this equipment item."""
        equipment = self.get_object()
        try:
            from .pdf_generator import generate_equipment_pdf
            pdf_buffer = generate_equipment_pdf(equipment)
            response = HttpResponse(
                pdf_buffer.read(), content_type='application/pdf')
            safe_name = equipment.equipment_number.replace(' ', '_')
            response[
                'Content-Disposition'] = f'attachment; filename="equipment_card_{safe_name}.pdf"'
            return response
        except Exception as e:
            return Response(
                {'detail': f'Error generating PDF: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get equipment statistics"""
        from django.db.models import Count, Q
        from datetime import date

        today = date.today()

        # Equipment is considered "not in validity" if:
        # - inspection_status is 'expired', OR
        # - next_inspection_date is in the past
        not_valid_q = Q(inspection_status='expired') | Q(
            next_inspection_date__lt=today)

        total = Equipment.objects.count()
        # "Active & valid" = status is active AND inspection is still valid
        active_valid = Equipment.objects.filter(
            status='active').exclude(not_valid_q).count()
        # "Not in validity" = active equipment whose inspection expired
        active_expired = Equipment.objects.filter(
            status='active').filter(not_valid_q).count()
        maintenance = Equipment.objects.filter(status='maintenance').count()
        inactive = Equipment.objects.filter(status='inactive').count()
        retired = Equipment.objects.filter(status='retired').count()

        stats = {
            'total': total,
            'by_type': dict(Equipment.objects.values('equipment_type').annotate(count=Count('id')).values_list('equipment_type', 'count')),
            'by_status': dict(Equipment.objects.values('status').annotate(count=Count('id')).values_list('status', 'count')),
            'active_valid': active_valid,
            'active_expired': active_expired,
            'not_active_total': active_expired + inactive + retired,
        }
        return Response(stats)

    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Export all equipment to CSV with current filters applied"""
        from django.http import HttpResponse

        # Apply filters to queryset
        queryset = self.filter_queryset(self.get_queryset())

        # Create CSV response
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="equipment_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'

        # Write BOM for Excel UTF-8 support
        response.write('\ufeff')

        writer = csv.writer(response)

        # Write headers (in Hebrew)
        headers = [
            'פריט ציוד', 'תחום ציוד', 'תחום על', 'סטטוס פריט ציוד', 'סטטוס בדיקות',
            'מספר סידורי פנימי', 'חברה', 'חברת שירות / קבלן',
            'אגף', 'חטיבה', 'מחלקה', 'תת מחלקה', 'יחידה',
            'מדינה', 'מחוז / איזור', 'עיר / יישוב', 'אתר / סניף',
            'מספר יא״מ', 'סטטוס אתר / סניף', 'קמפוס', 'כתובת',
            'מבנה / מתקן', 'קומה', 'חדר', 'מיקום', 'קו ייצור', 'פרויקט',
            'מספר רישיון / רישוי', 'יצרן', 'תאריך ייצור', 'מספר סידורי יצרן',
            'פקיעת תוקף אחריות', 'דגם', 'תאור', 'הערה', 'תגית',
            'אחראי/ת', 'בדיקות תקופתיות', 'חברת שירות', 'מספר קבצים',
            'ערכת ציוד', 'עובדים מוסמכים', 'עומס עבודה בטוח', 'לחץ מירבי מותר',
            'יחידת מדידה', 'רזולוציית מדידה', 'טווח מדידה',
            'תמונה', 'URL', 'GUID', 'נוצר על ידי', 'עודכן על ידי',
        ]
        writer.writerow(headers)

        # Write data
        for equipment in queryset:
            row = [
                equipment.equipment_number,
                equipment.get_equipment_type_display(),
                equipment.super_domain or '',
                equipment.get_status_display(),
                equipment.get_inspection_status_display() if equipment.inspection_status else '',
                equipment.internal_serial_number or '',
                equipment.employer or '',
                equipment.service_company or '',
                equipment.wing or '',
                equipment.division or '',
                equipment.department or '',
                equipment.sub_department or '',
                equipment.unit or '',
                equipment.country or '',
                equipment.district or '',
                equipment.city or '',
                equipment.site_name or '',
                equipment.yam_number or '',
                equipment.site_status or '',
                equipment.campus or '',
                equipment.address or '',
                equipment.building or '',
                equipment.floor_number or '',
                equipment.room or '',
                equipment.location_details or '',
                equipment.production_line or '',
                equipment.project or '',
                equipment.license_number or '',
                equipment.manufacturer or '',
                equipment.manufacture_date.strftime(
                    '%d/%m/%Y') if equipment.manufacture_date else '',
                equipment.serial_number or '',
                equipment.warranty_expiry.strftime(
                    '%d/%m/%Y') if equipment.warranty_expiry else '',
                equipment.model or '',
                equipment.description or '',
                equipment.notes or '',
                equipment.tag or '',
                equipment.inspector_name or '',
                equipment.periodic_inspections or '',
                equipment.service_company or '',
                str(equipment.file_count) if equipment.file_count else '0',
                equipment.equipment_set or '',
                equipment.certified_workers or '',
                equipment.safe_working_load or '',
                equipment.max_allowed_pressure or '',
                equipment.measurement_unit or '',
                equipment.measurement_resolution or '',
                equipment.measurement_range or '',
                equipment.image.url if equipment.image else '',
                equipment.url or '',
                str(equipment.id),
                equipment.created_by.username if equipment.created_by else '',
                equipment.updated_by.username if equipment.updated_by else '',
            ]
            writer.writerow(row)

        return response

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export all equipment to Excel (.xlsx) with current filters applied"""
        # Apply filters to queryset
        queryset = self.filter_queryset(self.get_queryset())

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Equipment"

        # Define headers (Hebrew)
        headers = [
            'פריט ציוד', 'תחום ציוד', 'תחום על', 'סטטוס פריט ציוד', 'סטטוס בדיקות',
            'מספר סידורי פנימי', 'חברה', 'חברת שירות / קבלן',
            'אגף', 'חטיבה', 'מחלקה', 'תת מחלקה', 'יחידה',
            'מדינה', 'מחוז / איזור', 'עיר / יישוב', 'אתר / סניף',
            'מספר יא״מ', 'סטטוס אתר / סניף', 'קמפוס', 'כתובת',
            'מבנה / מתקן', 'קומה', 'חדר', 'מיקום', 'קו ייצור', 'פרויקט',
            'מספר רישיון / רישוי', 'יצרן', 'תאריך ייצור', 'מספר סידורי יצרן',
            'פקיעת תוקף אחריות', 'דגם', 'תאור', 'הערה', 'תגית',
            'אחראי/ת', 'בדיקות תקופתיות', 'חברת שירות', 'מספר קבצים',
            'ערכת ציוד', 'עובדים מוסמכים', 'עומס עבודה בטוח', 'לחץ מירבי מותר',
            'יחידת מדידה', 'רזולוציית מדידה', 'טווח מדידה',
            'תמונה', 'URL', 'GUID', 'נוצר על ידי', 'עודכן על ידי',
        ]

        # Style header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data
        for row_num, equipment in enumerate(queryset, 2):
            col = 1

            def _w(val):
                nonlocal col
                ws.cell(row=row_num, column=col, value=val)
                col += 1

            _w(equipment.equipment_number)
            _w(equipment.get_equipment_type_display())
            _w(equipment.super_domain or '')
            _w(equipment.get_status_display())
            _w(equipment.get_inspection_status_display()
               if equipment.inspection_status else '')
            _w(equipment.internal_serial_number or '')
            _w(equipment.employer or '')
            _w(equipment.service_company or '')
            _w(equipment.wing or '')
            _w(equipment.division or '')
            _w(equipment.department or '')
            _w(equipment.sub_department or '')
            _w(equipment.unit or '')
            _w(equipment.country or '')
            _w(equipment.district or '')
            _w(equipment.city or '')
            _w(equipment.site_name or '')
            _w(equipment.yam_number or '')
            _w(equipment.site_status or '')
            _w(equipment.campus or '')
            _w(equipment.address or '')
            _w(equipment.building or '')
            _w(equipment.floor_number or '')
            _w(equipment.room or '')
            _w(equipment.location_details or '')
            _w(equipment.production_line or '')
            _w(equipment.project or '')
            _w(equipment.license_number or '')
            _w(equipment.manufacturer or '')
            _w(equipment.manufacture_date.strftime('%d/%m/%Y')
               if equipment.manufacture_date else '')
            _w(equipment.serial_number or '')
            _w(equipment.warranty_expiry.strftime('%d/%m/%Y')
               if equipment.warranty_expiry else '')
            _w(equipment.model or '')
            _w(equipment.description or '')
            _w(equipment.notes or '')
            _w(equipment.tag or '')
            _w(equipment.inspector_name or '')
            _w(equipment.periodic_inspections or '')
            _w(equipment.service_company or '')
            _w(equipment.file_count or 0)
            _w(equipment.equipment_set or '')
            _w(equipment.certified_workers or '')
            _w(equipment.safe_working_load or '')
            _w(equipment.max_allowed_pressure or '')
            _w(equipment.measurement_unit or '')
            _w(equipment.measurement_resolution or '')
            _w(equipment.measurement_range or '')
            _w(equipment.image.url if equipment.image else '')
            _w(equipment.url or '')
            _w(str(equipment.id))
            _w(equipment.created_by.username if equipment.created_by else '')
            _w(equipment.updated_by.username if equipment.updated_by else '')

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="equipment_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['post'])
    def import_csv(self, request):
        """Import equipment from CSV file.

        Supports both the legacy positional CSV format and header-based CSV files.
        Header-based import recognizes common Hebrew headers (incl. the set used by
        the /equipment/import UI).
        """
        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        csv_file = request.FILES['file']

        # Detect encoding – prefer Hebrew-specific encodings for
        # files that chardet may misidentify as Windows-1252.
        raw_data = csv_file.read()
        result = chardet.detect(raw_data)
        encoding = result['encoding']

        # Decode file – try Hebrew encodings first when chardet is unsure
        decoded_file = None
        if encoding and encoding.lower() in ('windows-1252', 'iso-8859-1', 'latin-1', 'ascii'):
            # chardet often misdetects Hebrew (cp1255) as Windows-1252;
            # try Hebrew encodings first.
            for hebrew_enc in ('cp1255', 'windows-1255', 'iso-8859-8'):
                try:
                    candidate = raw_data.decode(hebrew_enc)
                    # Quick sanity: Hebrew letters live in \u0590-\u05FF
                    if any('\u0590' <= ch <= '\u05FF' for ch in candidate[:2000]):
                        decoded_file = candidate
                        break
                except Exception:
                    continue
        if decoded_file is None:
            try:
                decoded_file = raw_data.decode(encoding)
            except Exception:
                decoded_file = raw_data.decode('utf-8', errors='replace')

        # Detect delimiter (fallback to ';')
        sample = decoded_file[:4096]
        delimiter = ';'
        try:
            dialect = csv.Sniffer().sniff(
                sample, delimiters=[',', ';', '\t', '|'])
            delimiter = dialect.delimiter
        except Exception:
            if sample.count(';') >= sample.count(','):
                delimiter = ';'
            elif sample.count(',') > 0:
                delimiter = ','

        csv_reader = csv.DictReader(io.StringIO(
            decoded_file), delimiter=delimiter)
        if not csv_reader.fieldnames:
            return Response({'error': 'CSV file has no header row'}, status=status.HTTP_400_BAD_REQUEST)

        raw_headers = list(csv_reader.fieldnames)
        header_map = {raw: _normalize_header(raw) for raw in raw_headers}

        # -----------------------------------------------------------
        # Check whether ANY header matches a known candidate.
        # If none match, fall back to positional column mapping so
        # that garbled-header CSVs can still be imported.
        # -----------------------------------------------------------
        normalized_header_set = set(header_map.values())
        all_known_candidates = set()
        for candidates in HEADER_CANDIDATES.values():
            for c in candidates:
                all_known_candidates.add(_normalize_header(c))

        headers_recognized = bool(normalized_header_set & all_known_candidates)

        # Positional mapping for the well-known 17-column CSV layout
        # exported from the equipment management system.
        POSITIONAL_MAP = {
            0: '',                  # row number – ignored
            1: 'תחום ציוד',        # equipment_type
            2: 'תחום על',          # super_domain
            3: 'סטטוס פריט ציוד',  # status
            4: 'סטטוס בדיקות',     # inspection_status
            5: 'מספר סידורי פנימי',  # equipment_number (internal serial)
            6: 'חברה',             # company / employer
            7: 'מחלקה',            # department
            8: 'אתר / סניף',      # site_name
            9: 'מיקום',            # location
            10: 'יצרן',           # manufacturer
            11: 'דגם',            # model
            12: 'מספר סידורי יצרן',  # serial_number
            13: 'תאור',           # description
            14: 'הערה',           # notes
            15: 'בודק/ת',         # inspector
            16: 'בדיקות תקופתיות',  # periodic_inspections
        }

        def rows_iter():
            # Re-read the file for positional fallback when needed
            if not headers_recognized and len(raw_headers) >= 16:
                # Positional fallback – build normalized rows using the
                # known column-index → Hebrew-header mapping.
                pos_reader = csv.reader(
                    io.StringIO(decoded_file), delimiter=delimiter)
                next(pos_reader, None)  # skip header
                for row_index, values in enumerate(pos_reader, start=2):
                    normalized_row = {}
                    for col_idx, hdr_name in POSITIONAL_MAP.items():
                        if col_idx < len(values) and hdr_name:
                            normalized_row[_normalize_header(
                                hdr_name)] = values[col_idx]
                    if not normalized_row:
                        continue
                    yield (row_index, normalized_row)
            else:
                # Header-based mapping (normal path)
                for row_index, row in enumerate(csv_reader, start=2):
                    normalized_row = {}
                    for raw_key, raw_value in row.items():
                        normalized_key = header_map.get(
                            raw_key, _normalize_header(raw_key))
                        normalized_row[normalized_key] = raw_value
                    yield (row_index, normalized_row)

        result = _import_normalized_rows(
            user=request.user, rows_iter=rows_iter())
        return Response(result)

    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        """Import equipment from an Excel .xlsx file (first sheet)."""
        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        excel_file = request.FILES['file']
        filename = getattr(excel_file, 'name', '') or ''
        if not filename.lower().endswith('.xlsx'):
            return Response({'error': 'Only .xlsx files are supported'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            from io import BytesIO
            content = excel_file.read()
            wb = load_workbook(filename=BytesIO(content), data_only=True)
            ws = wb.active

            rows = ws.iter_rows(values_only=True)
            header_row = next(rows, None)
            if not header_row:
                return Response({'error': 'Excel file is empty'}, status=status.HTTP_400_BAD_REQUEST)

            raw_headers = [_normalize_header(h) for h in header_row]
            if all(h == '' for h in raw_headers):
                return Response({'error': 'Excel header row is empty'}, status=status.HTTP_400_BAD_REQUEST)

            def rows_iter():
                row_index = 2
                for values in rows:
                    if values is None:
                        row_index += 1
                        continue

                    # Build row dict by headers; trim trailing empty cells
                    normalized_row = {}
                    for col_index, header in enumerate(raw_headers):
                        if not header:
                            continue
                        cell_value = values[col_index] if col_index < len(
                            values) else None
                        if cell_value is None:
                            continue
                        normalized_row[header] = str(cell_value).strip() if not isinstance(
                            cell_value, (int, float)) else str(cell_value)

                    # Skip blank lines
                    if len(normalized_row) == 0:
                        row_index += 1
                        continue

                    yield (row_index, normalized_row)
                    row_index += 1

            result = _import_normalized_rows(
                user=request.user, rows_iter=rows_iter())
            return Response(result)

        except Exception as e:
            return Response({'error': f'Failed to import Excel: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
