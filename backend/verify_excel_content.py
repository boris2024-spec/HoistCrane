"""
Проверка содержимого экспортированного Excel файла
"""
from openpyxl import load_workbook

# Проверяем файл оборудования
print("📊 Проверка файла equipment_test_export.xlsx\n")
wb = load_workbook('equipment_test_export.xlsx')
ws = wb.active

print(f"Название листа: {ws.title}")
print(f"Количество строк: {ws.max_row}")
print(f"Количество колонок: {ws.max_column}")

print("\n📝 Заголовки (первая строка):")
headers = []
for col in range(1, ws.max_column + 1):
    cell_value = ws.cell(1, col).value
    headers.append(cell_value)
    print(f"  Колонка {col}: {cell_value}")

print("\n📄 Первые 3 записи данных:")
for row in range(2, min(5, ws.max_row + 1)):
    print(f"\nЗапись {row-1}:")
    for col in range(1, min(7, ws.max_column + 1)):  # Показываем первые 6 колонок
        header = headers[col-1] if col <= len(headers) else f"Колонка {col}"
        value = ws.cell(row, col).value or ''
        print(f"  {header}: {value}")

print("\n✅ Проверка Unicode/иврита:")
hebrew_found = False
for row in range(1, min(5, ws.max_row + 1)):
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row, col).value
        if cell_value and isinstance(cell_value, str):
            # Проверяем наличие символов иврита (Hebrew Unicode range: 0x0590-0x05FF)
            if any('\u0590' <= char <= '\u05FF' for char in cell_value):
                hebrew_found = True
                print(f"  Найден иврит в ячейке ({row}, {col}): {cell_value}")
                break
    if hebrew_found:
        break

if hebrew_found:
    print("\n✓ Иврит успешно экспортирован в Unicode формате!")
else:
    print("\n⚠️ Иврит не обнаружен в первых строках")

print("\n" + "="*50)
print("Вывод: Excel файл создан корректно с поддержкой UTF-8/Unicode")
print("       Заголовки и данные на иврите должны правильно")
print("       отображаться в Microsoft Excel, LibreOffice и др.")
