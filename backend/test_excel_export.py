"""
Простой тест экспорта Excel с поддержкой иврита
"""
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook
import os
import sys
import django

# Setup Django environment
sys.path.append(os.path.dirname(__file__))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'hoistcraneproject.settings')
django.setup()


# Создаем тестовый Excel файл с ивритом
wb = Workbook()
ws = wb.active
ws.title = "Test Hebrew"

# Заголовки на иврите
headers = [
    'פריט ציוד', 'סוג ציוד', 'סטטוס', 'יצרן', 'דגם', 'מספר סידורי',
]

# Стилизуем заголовки
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4",
                          end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# Добавляем тестовые данные с ивритом
test_data = [
    ['EQ-001', 'מנוף', 'פעיל', 'יצרן ישראלי', 'דגם A', 'SN-12345'],
    ['EQ-002', 'מנוף רמה', 'תחזוקה', 'יצרן אחר', 'דגם B', 'SN-67890'],
]

for row_num, row_data in enumerate(test_data, 2):
    for col_num, value in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col_num, value=value)

# Автоматическая ширина колонок
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column_letter].width = adjusted_width

# Сохраняем файл
output_file = f'test_hebrew_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
wb.save(output_file)

print(f"✓ Тестовый файл успешно создан: {output_file}")
print(f"✓ Файл содержит заголовки и данные на иврите")
print(f"✓ Откройте файл в Excel для проверки правильного отображения иврита")
