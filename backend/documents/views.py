from rest_framework import viewsets, filters, permissions
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
from django_filters.rest_framework import DjangoFilterBackend
from django_filters import rest_framework as django_filters
from .models import Document
from .serializers import DocumentSerializer, DocumentUploadSerializer
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse


class DocumentFilter(django_filters.FilterSet):
    document_type = django_filters.ChoiceFilter(
        choices=Document.DOCUMENT_TYPE_CHOICES)
    equipment = django_filters.UUIDFilter(field_name='equipment__id')
    uploaded_from = django_filters.DateFilter(
        field_name='uploaded_at', lookup_expr='gte')
    uploaded_to = django_filters.DateFilter(
        field_name='uploaded_at', lookup_expr='lte')

    class Meta:
        model = Document
        fields = ['document_type', 'equipment']


class DocumentViewSet(viewsets.ModelViewSet):
    queryset = Document.objects.select_related('equipment', 'uploaded_by')
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    filter_backends = [DjangoFilterBackend,
                       filters.SearchFilter, filters.OrderingFilter]
    filterset_class = DocumentFilter
    search_fields = ['title', 'description', 'equipment__equipment_number']
    ordering_fields = ['uploaded_at', 'document_date', 'title']
    ordering = ['-uploaded_at']

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return DocumentUploadSerializer
        return DocumentSerializer

    def perform_create(self, serializer):
        serializer.save(uploaded_by=self.request.user)

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export all documents to Excel (.xlsx) with current filters applied"""
        # Apply filters to queryset
        queryset = self.filter_queryset(self.get_queryset())

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Documents"

        # Define headers (Hebrew)
        headers = [
            'מזהה', 'כותרת', 'סוג מסמך', 'ציוד', 'תיאור',
            'תאריך מסמך', 'תאריך תפוגה', 'גודל קובץ (KB)',
            'הועלה על ידי', 'תאריך העלאה'
        ]

        # Style header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data
        for row_num, doc in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=str(doc.id))
            ws.cell(row=row_num, column=2, value=doc.title)
            ws.cell(row=row_num, column=3,
                    value=doc.get_document_type_display())
            ws.cell(row=row_num, column=4,
                    value=doc.equipment.equipment_number if doc.equipment else '')
            ws.cell(row=row_num, column=5, value=doc.description or '')
            ws.cell(row=row_num, column=6, value=doc.document_date.strftime(
                '%d/%m/%Y') if doc.document_date else '')
            ws.cell(row=row_num, column=7, value=doc.expiry_date.strftime(
                '%d/%m/%Y') if doc.expiry_date else '')
            ws.cell(row=row_num, column=8, value=round(
                doc.file_size / 1024, 2) if doc.file_size else '')
            ws.cell(row=row_num, column=9,
                    value=doc.uploaded_by.username if doc.uploaded_by else '')
            ws.cell(row=row_num, column=10, value=doc.uploaded_at.strftime(
                '%d/%m/%Y %H:%M') if doc.uploaded_at else '')

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="documents_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response
